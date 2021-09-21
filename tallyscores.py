from openpyxl.styles import PatternFill, Font

def color_fill(ws, score, row_num):
  line = ws.cell(row=row_num, column=2).value.split(' ')
  score_split = score.split(' ')

  # cycle through four pick columns; odd = common, even = spread
  for col in range(5, 9):
    pick_cell = ws.cell(row=row_num, column=col)

    # common pick
    if col % 2 == 1:
      pick_cell_split = pick_cell.value.split(' ')

      # if pick team == winning team, color cell green
      if pick_cell_split[0].upper() == score_split[0]:
        pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        spread = score_split[1].split('-')

        # if game spread == pick spread, spread font == yellow
        if int(spread[0]) - int(spread[1]) == int(pick_cell_split[2]):
          pick_cell.font = Font(color="FFFB00") # yellow
      # else color cell red    
      else:
        pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red

    # TODO - against line
    else:
      continue