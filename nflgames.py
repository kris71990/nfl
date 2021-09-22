# nflgames.py 
# gets nfl weekly matchups and enters them into spreadsheet along with
# team records from nflteams.py

import requests, bs4, os, sys, re, nflteams, byeteams, teaminfo, weekInfo
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from dotenv import load_dotenv
load_dotenv()

week_num = sys.argv[1]

if 'Week' in week_num:
  week_num = week_num.split(' ')[1]

# get weekly matchup data
url = 'http://www.vegasinsider.com/nfl/odds/las-vegas/'
res = requests.get(url)
res.raise_for_status()
soup = bs4.BeautifulSoup(res.text, 'html.parser')

# find and parse soup for game matchups
def findMatchups(byes):
  teams_html = soup.select('a[class=tabletext]')
  teams_raw = []

  for item in teams_html:
    team = item.getText()
    teams_raw.append(team)

  # cut list down to accomodate extra games website is now displaying
  total_active_teams = 32 - len(byes) 
  teams_raw = teams_raw[:total_active_teams]

  # create the matchups from teams_raw list
  matchups = []
  matchup_num = 0
  matchup_num_2 = matchup_num + 1
  num_games = int(len(teams_raw) / 2)
  dict = nflteams.team_records_dict 

  for i in range(0, num_games):
    matchups.append(teams_raw[matchup_num] + ' ' + dict.get(teams_raw[matchup_num]) + \
    ' at ' + teams_raw[matchup_num_2] + ' ' + dict.get(teams_raw[matchup_num_2]))
    matchup_num += 2
    matchup_num_2 += 2

  return matchups

# finds odds soup and parses it to find the VI consensus for every matchup
def findOdds():
  odds_html = soup.findAll('td', class_=['cellTextNorm', 'cellTextHot'])

  odds = []

  for x in range(8, len(odds_html), 9):
    game_odd_text = odds_html[x].find('br').getText()
    game_odd = game_odd_text.replace('\t', '').replace('\n', '').replace('\xa0', '')

    # find if home or road team is favored
    if (len(game_odd.split('u')) > 1):
      team_favored_odd = {}

      if ('PK' in game_odd): # neither
        team_favored_odd['n'] = 'Pick'
        odds.append(team_favored_odd)
        continue

      if (game_odd.startswith('-')): # road
        parsed_line = game_odd.split('-', 2)[1]
        team_favored_odd['r'] = '-' + re.split('EV', parsed_line)[0]
        odds.append(team_favored_odd)
      else: # home
        parsed_line = game_odd.split('-10', 1)[1].split('-')[1]
        team_favored_odd['h'] = '-' + re.split('EV', parsed_line)[0]
        odds.append(team_favored_odd)

  return odds

byeData = byeteams.get_bye_teams(week_num)
formattedByes = byeteams.formatByes(byeData)
matchups = findMatchups(byeData)
odds = findOdds()

# creates dictionary from matchup/odds data 
def createGameData():
  data = {}
  keys = range(len(matchups))

  for i in keys:
    data[i] = { matchups[i] : None }

  for i in keys:
    teams = matchups[i].split(' at ')
    rx = re.compile(r'([a-zA-Z\s.-]+)', re.I)

    if ('h' in odds[i]): # if home team is favoured
      stripped_team = rx.search(teams[1]).group(0).strip()
      team_abbreviation = teaminfo.abbreviations[stripped_team]
      line = '%s %s' % (team_abbreviation, odds[i]['h'])
    elif ('r' in odds[i]): # if road team is favoured
      stripped_team = rx.search(teams[0]).group(0).strip()
      team_abbreviation = teaminfo.abbreviations[stripped_team]
      line = '%s %s' % (team_abbreviation, odds[i]['r'])
    else:
      line = 'Pick'

    data[i][matchups[i]] = line
    print(line)

  print('\n')
  return data

# Print all weekly game information to console
print('\nGames for Week %s\n' % (week_num))
gameData = createGameData()
print(gameData)
print('\n')
print(formattedByes)

# open spreadsheet and write info
def write_game_info():
  print('\nWriting to spreadsheet...\n')

  os.chdir(os.getenv('LOCATION'))
  wb = load_workbook(os.getenv('EXCEL_FILE'))
  sheet = wb.get_sheet_by_name('Sheet 1')

  #find spreadsheet start row and write game info to appropriate cells
  write_matchup_num = 0
  start_row = 0
  if 'Week' not in week_num:
    if (int(week_num) > 17):
      start_week = weekInfo.playoff_week_titles[week_num]
    else:
      start_week = 'Week ' + week_num
  else:
    start_week = week_num
  bye_row = 0

  for cell in sheet.columns[0]:
    if cell.value == start_week:
      start_row += 2
      for row_num in range(start_row, len(matchups) + start_row):
        game = sheet.cell(row=row_num, column=1)

        if (row_num == (len(matchups) + start_row) - 1):
          bye_row = row_num

        game.value = matchups[write_matchup_num]

        line = sheet.cell(row=row_num, column=2)
        line.value = gameData[write_matchup_num][matchups[write_matchup_num]]
        write_matchup_num += 1       
    else: 
      start_row += 1
  
  bye = sheet.cell(row=bye_row+1, column=1)
  if formattedByes is not None:
    bye.value = formattedByes

  week = sheet.cell(row=bye_row+1, column=3)
  week.value = 'Week =>'
  week.alignment = Alignment(horizontal='right')
  week.font = Font(italic=True)

  total = sheet.cell(row=bye_row+2, column=3)
  total.value = 'Total =>'
  total.alignment = Alignment(horizontal='right')
  week.font = Font(italic=True)

  wb.save(os.getenv('EXCEL_FILE_NEW'))
  print('Done')

write_game_info()