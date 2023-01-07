# nflgames.py 
# gets nfl weekly matchups and enters them into spreadsheet along with
# team records from nflteams.py

import re, schedule.nflteams, schedule.byeteams
from assets import weekInfo, teaminfo, soup
from openpyxl.styles import Alignment, Font, PatternFill

# find and parse soup for game matchups
def find_matchups(byes, soup):
  teams_html = soup.findAll('div', class_=['all-markets-team-stats-box'])
  teams_raw = []

  for item in teams_html:
    team = item.getText().strip()
    team_without_name = teaminfo.reformat_without_nickname(team)
    teams_raw.append(team_without_name)

  # cut list down to accomodate extra games website is now displaying
  bye_length = 0 if byes is None else len(byes)
  total_active_teams = 32 - bye_length 
  teams_raw = teams_raw[:total_active_teams]

  # create the matchups from teams_raw list
  matchups = []
  matchup_num = 0
  matchup_num_2 = matchup_num + 1
  num_games = int(len(teams_raw) / 2)
  records = schedule.nflteams.get_team_records()

  for i in range(0, num_games):
    matchups.append(teams_raw[matchup_num] + ' (' + records.get(teams_raw[matchup_num]) + ') at ' + teams_raw[matchup_num_2] + ' (' + records.get(teams_raw[matchup_num_2]) + ')')
    matchup_num += 2
    matchup_num_2 += 2

  return matchups

# finds odds soup and parses it to find the VI consensus for every matchup
def find_odds(soup):
  odds_html = soup.find_all('div', class_='odds-box')
  odds = []

  for x in range(1, len(odds_html), 60):
    game_odd = odds_html[x].contents[1].get_text().strip()

    # find if home or road team is favored, toggle location variable after every cycle to track home/road odds
    if (game_odd.startswith('-')):
      location_favored = 'a'
      odds.append({ location_favored: game_odd })
    else:
      location_favored = 'h'
      odds.append({ location_favored: '-' + game_odd })

    # if ('PK' in game_odd): # neither
    #   team_favored_odd['n'] = 'Pick'
    #   odds.append(team_favored_odd)
    #   continue

  return odds

# creates dictionary from matchup/odds data 
def create_game_data(matchups, odds):
  data = {}
  keys = range(len(matchups))


  for i in keys:
    data[i] = { matchups[i] : None }

  for i in keys:
    teams = matchups[i].split(' at ')
    rx = re.compile(r'([a-zA-Z\s.-]+)', re.I)

    if ('h' in odds[i]): # if home team is favoured
      stripped_team = rx.search(teams[1]).group(0).strip()
      team_abbreviation = teaminfo.abbreviations[stripped_team]
      line = '%s %s' % (team_abbreviation, odds[i]['h'])
    elif ('a' in odds[i]): # if road team is favoured
      stripped_team = rx.search(teams[0]).group(0).strip()
      team_abbreviation = teaminfo.abbreviations[stripped_team]
      line = '%s %s' % (team_abbreviation, odds[i]['a'])
    else:
      line = 'Pick'

    data[i][matchups[i]] = line
  return data

def write_footer_header(ss, footer_row, week):
  # blue color fill footer
  for row in ss['sheet'].iter_rows(min_row=footer_row, max_row=footer_row):
    for cell in row:
      cell.fill = PatternFill("solid", fgColor="43889D")
  
  # next week header    
  week_num = int(week)
  if (week_num + 1 < 19):
    next_week = 'Week ' + str(week_num + 1)
  elif (week_num + 1 < 23):
    next_week = weekInfo.playoff_week_titles[str(week_num + 1)]
  else:
    # super bowl
    return
  
  next_week_cell = ss['sheet'].cell(row=footer_row + 1, column=1)
  next_week_cell.value = next_week
  next_week_cell.font = Font(name='Times New Roman', size=14, bold=True)
  next_week_cell.alignment = Alignment(horizontal='center')
  return

# open spreadsheet and write info
def write_game_info(ss, week, matchups, game_data, formatted_byes):
  print('\nWriting to spreadsheet...\n')

  #find spreadsheet start row and write game info to appropriate cells
  write_matchup_num = 0
  start_row = 0
  if (int(week) > 18):
    start_week = weekInfo.playoff_week_titles[week]
  else:
    start_week = 'Week ' + week
  bye_row = 0

  for cell in ss['sheet']['A']:
    if cell.value == start_week:
      start_row += 2
      for row_num in range(start_row, len(matchups) + start_row):
        game = ss['sheet'].cell(row=row_num, column=1)

        if (row_num == (len(matchups) + start_row) - 1):
          bye_row = row_num

        game.value = matchups[write_matchup_num]
        game.alignment = Alignment(vertical='bottom', horizontal='left')
        game.font = Font(name='Times New Roman', size=12, bold=True)

        line = ss['sheet'].cell(row=row_num, column=2)
        line.value = game_data[write_matchup_num][matchups[write_matchup_num]]
        line.alignment = Alignment(vertical='bottom', horizontal='center')
        line.font = Font(name='Times New Roman', size=12)
        write_matchup_num += 1       
    else: 
      start_row += 1
  
  bye = ss['sheet'].cell(row=bye_row + 1, column=1)
  if formatted_byes is not None:
    bye.value = formatted_byes

  week_cell = ss['sheet'].cell(row=bye_row + 1, column=3)
  week_cell.value = 'Week =>'
  week_cell.alignment = Alignment(horizontal='right')
  week_cell.font = Font(name='Times New Roman', size=12, italic=True)

  total_cell = ss['sheet'].cell(row=bye_row + 2, column=3)
  total_cell.value = 'Total =>'
  total_cell.alignment = Alignment(horizontal='right')
  total_cell.font = Font(name='Times New Roman', size=12, italic=True)

  footer_row = bye_row + 3
  write_footer_header(ss, footer_row, week)
  return

def printable_game_data(game_data):
  for key, value in game_data.items():
    for game, line in value.items():
      x = '{0} => {1}'.format(game, line)
      print(x)

def init(ss, week):
  odds_soup = soup.get_odds_soup()
  bye_data = schedule.byeteams.get_bye_teams(week)
  formatted_byes = schedule.byeteams.format_byes(bye_data)
  matchups = find_matchups(bye_data, odds_soup)
  odds = find_odds(odds_soup)

  # Print all weekly game information to console
  if int(week) > 18:
    print('\n%s\n' % (weekInfo.playoff_week_titles[week]))
  else:
    print('\nGames for Week %s\n' % (week))

  game_data = create_game_data(matchups, odds)
  printable_game_data(game_data)
  print('\n')

  if formatted_byes is None:
    print('No Byes')
  else:
    print(formatted_byes)
    
  write_game_info(ss, week, matchups, game_data, formatted_byes)
  return
