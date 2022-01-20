# nflscores.py 
# gets nfl weekly scores and enters them into spreadsheet 

import re, results.tallyscores
from assets import teaminfo, weekInfo, soup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def get_scores(week):
  espn_format = soup.convert_espn(week)
  scores_soup = soup.get_scores_soup(espn_format['espn_week'], espn_format['espn_season_type'])
  scores_raw = scores_soup.find_all('a', { 'name': '&lpos=nfl:schedule:score' })
  scores = {}

  for each in scores_raw:
    score = each.get_text()
    print(score)
    score_split = score.split(' ')
    # if tie *** may be buggy ***
    if (score_split[1].strip(', ') == score_split[3]):
      scores[score_split[0] + '-' + score_split[2]] = '%s-%s' % (score_split[1].strip(','), score_split[3])
    else:
      # if overtime, add OT to string
      if (len(score_split) > 4):
        scores[score_split[0]] = '%s-%s %s' % (score_split[1].strip(','), score_split[3], score_split[4])
      else:
        scores[score_split[0]] = '%s-%s' % (score_split[1].strip(','), score_split[3])

  return scores

def write_scores(ss, week):
  print('\nWriting to spreadsheet...\n')

  if (int(week) > 18):
    print('%s Playoff results:\n' % weekInfo.playoff_week_titles[week])
  else:
    print('Week %s results:\n' % week)
    
  scores = get_scores(week)

  # find spreadsheet start row and write scores to appropriate cells
  write_score_index = 0
  start_row = 0
  if int(week) > 18:
    start_week = weekInfo.playoff_week_titles[str(int(week))]
  else:
    start_week = 'Week ' + week

  for cell in ss['sheet']['A']:
    if cell.value == start_week:
      start_row += 2
      for row_num in range(start_row, len(scores) + start_row):
        score_cell = ss['sheet'].cell(row=row_num, column=3)
        matchup_cell = ss['sheet'].cell(row=row_num, column=1)
        matchup_cell_text = matchup_cell.value

        rx = re.compile(r'([a-zA-Z\s.]+)', re.I)
        teams_raw = rx.findall(matchup_cell_text)
        teams = [x.strip(' ') for x in teams_raw]
        team1abbr = teaminfo.abbreviations[teams[0]].upper()
        team2abbr = teaminfo.abbreviations[teams[1][3:]].upper()
        
        if (team1abbr in scores):
          score = '%s %s' % (team1abbr, scores[team1abbr])
          score_cell.value = score
          score_cell.font = Font(name='Times New Roman', size=12)
          score_cell.alignment = Alignment(horizontal='center', vertical='center')
        elif (team2abbr in scores):
          score = '%s %s' % (team2abbr, scores[team2abbr])
          score_cell.value = score
          score_cell.font = Font(name='Times New Roman', size=12)
          score_cell.alignment = Alignment(horizontal='center', vertical='center')
        # tie
        else:
          score = 'TIE ' + scores[team1abbr + '-' + team2abbr]
          score_cell.value = score
          score_cell.font = Font(name='Times New Roman', size=12)
          score_cell.alignment = Alignment(horizontal='center', vertical='center')
        results.tallyscores.color_fill(ss['sheet'], score, row_num)
        write_score_index += 1       
    else: 
      start_row += 1
  return
