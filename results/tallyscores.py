from openpyxl.styles import PatternFill, Font

def color_fill(ws, score, row_num):
  score_split = score.split(' ')

  # cycle through four pick columns; odd = common, even = spread
  for col in range(5, 9):
    pick_cell = ws.cell(row=row_num, column=col)
    pick_cell_text = pick_cell.value

    # common pick
    if col % 2 == 1:
      pick_cell_split = pick_cell_text.split(' ')

      # if pick team == winning team, color cell green
      if pick_cell_split[0].upper() == score_split[0]:
        pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        spread = score_split[1].split('-')
        
        # if game spread == pick spread, spread font == bold yellow
        if int(spread[0]) - int(spread[1]) == int(pick_cell_split[2]):
          pick_cell.font = Font(name='Times New Roman', size=12, color="FFFB00", bold=True)
      # else color cell red
      else: 
        pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
    # against line
    else:
      line = ws.cell(row=row_num, column=2).value.upper()
      line_list = line.split(' -')
      spread = score_split[1].split('-')

      # if Pick'em
      # Honest question: what happens if a pickem game ends in a tie????
      if line == 'PICK':
        team = score_split[0]
        if team == pick_cell_text.upper():
          pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        else:
          pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
      else:
        try:
          line_num = int(line_list[1])
        except ValueError:
          line_num = float(line_list[1][:-1] + '.5')

        if score_split[0] == line_list[0] and (int(spread[0]) - int(spread[1])) >= line_num:
          team = score_split[0]
          if team == pick_cell_text.upper():
            pick_cell.fill = PatternFill("solid", fgColor="009051") # green
          else:
            pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
        # if tie, line pick is green if opposite of favored team
        elif score_split[0] == 'TIE':
          if line_list[0] == pick_cell_text.upper():
            pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
          else:
            pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        elif score_split[0] != line_list[0]:
          team = score_split[0]
          if team == pick_cell_text.upper():
            pick_cell.fill = PatternFill("solid", fgColor="009051") # green
          else: 
            pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
        else:
          if pick_cell_text.upper() != line_list[0]:
            pick_cell.fill = PatternFill("solid", fgColor="009051") # green
          else:
            pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
